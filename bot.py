import asyncio
import logging
import os
from dotenv import load_dotenv
from datetime import datetime

from aiogram import Bot, Dispatcher, F
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
    FSInputFile
)
from aiogram.filters import Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage

from openpyxl import Workbook, load_workbook
load_dotenv()

TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = os.getenv("ADMIN_ID")
EXCEL_FILE = "zapisi.xlsx"

logging.basicConfig(level=logging.INFO)

bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())


# ---------- Excel ----------

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Дата обновления",
            "User ID",
            "Возраст",
            "ФИО ребенка",
            "ФИО родителя",
            "Телефон",
            "Статус",
            "Напомнено ПТ",
            "Напомнено СБ"
        ])
        wb.save(EXCEL_FILE)


def save_or_update_excel(user_id, data, status):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    found = False

    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(user_id):

            row[0].value = datetime.now().strftime("%Y-%m-%d %H:%M")
            row[2].value = data.get("age", "")
            row[3].value = data.get("child_name", "")
            row[4].value = data.get("parent_name", "")
            row[5].value = data.get("phone", "")
            row[6].value = status

            found = True
            break

    if not found:
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            user_id,
            data.get("age", ""),
            data.get("child_name", ""),
            data.get("parent_name", ""),
            data.get("phone", ""),
            status,
            "Нет",
            "Нет"
        ])

    wb.save(EXCEL_FILE)


def update_status(user_id, status):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(user_id):
            row[6].value = status
            row[0].value = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb.save(EXCEL_FILE)


def get_active_users():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    users = []

    for row in ws.iter_rows(min_row=2):
        if row[6].value == "Подтвержден":
            users.append(row)

    return users


# ---------- Клавиатуры ----------

main_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Записаться")],
        [KeyboardButton(text="Отменить запись")]
    ],
    resize_keyboard=True
)

age_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="6-8 лет")],
        [KeyboardButton(text="9-11 лет")],
        [KeyboardButton(text="12-14 лет")]
    ],
    resize_keyboard=True
)


def admin_confirm_kb(user_id):
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="Подтвердить",
                    callback_data=f"confirm_{user_id}"
                ),
                InlineKeyboardButton(
                    text="Отклонить",
                    callback_data=f"reject_{user_id}"
                )
            ]
        ]
    )


admin_panel = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Активные записи")],
        [KeyboardButton(text="Скачать Excel")]
    ],
    resize_keyboard=True
)


# ---------- Состояния ----------

class Form(StatesGroup):
    age = State()
    child_name = State()
    parent_name = State()
    phone = State()


# ---------- Антиспам ----------

user_last_message = {}

def anti_spam(user_id):
    now = datetime.now().timestamp()

    if user_id in user_last_message:
        if now - user_last_message[user_id] < 1:
            return False

    user_last_message[user_id] = now
    return True


# ---------- Старт ----------

@dp.message(Command("start"))
async def start(message: Message):
    if message.from_user.id == ADMIN_ID:
        await message.answer("Админ панель", reply_markup=admin_panel)
    else:
        await message.answer("""
В Альметьевске на этой неделе пройдет бесплатный мастер-класс для детей 6–14 лет по работе с нейросетями! 💻

✅ Ребенок познакомится с магией нейросетей и узнает, как применять искусственный интеллект в деле

✅ Создаст своего персонажа для игры с помощью искусственного интеллекта

✅ Запрограммирует героя Minecraft

✅ Длительность мастер-класса — 60 минут  
Ничего брать с собой не надо

Текущий уровень не важен — всему научим 👌🏻

Мастер-классы проходят по адресу:  
📍 ул. Гагарина, дом 3, г. Альметьевск

Чтобы записаться на бесплатный мастер-класс, выберете "Записаться" в меню ниже 
""",
            reply_markup=main_kb
        )


# ---------- Запись ----------

@dp.message(F.text == "Записаться")
async def signup(message: Message, state: FSMContext):
    if not anti_spam(message.from_user.id):
        return

    await message.answer("Выберите возраст своего ребенка:", reply_markup=age_kb)
    await state.set_state(Form.age)


@dp.message(Form.age)
async def age(message: Message, state: FSMContext):
    await state.update_data(age=message.text)
    await message.answer("Введите ФИО ребенка:")
    await state.set_state(Form.child_name)


@dp.message(Form.child_name)
async def child(message: Message, state: FSMContext):
    await state.update_data(child_name=message.text)
    await message.answer("Введите ФИО родителя:")
    await state.set_state(Form.parent_name)


@dp.message(Form.parent_name)
async def parent(message: Message, state: FSMContext):
    await state.update_data(parent_name=message.text)
    await message.answer("Введите номер телефона:")
    await state.set_state(Form.phone)


@dp.message(Form.phone)
async def phone(message: Message, state: FSMContext):
    data = await state.get_data()
    data["phone"] = message.text

    user_id = message.from_user.id

    save_or_update_excel(user_id, data, "Ожидает подтверждения")

    text = (
        f"Новая запись!\n\n"
        f"Возраст: {data['age']}\n"
        f"Ребенок: {data['child_name']}\n"
        f"Родитель: {data['parent_name']}\n"
        f"Телефон: {data['phone']}"
    )

    await bot.send_message(
        ADMIN_ID,
        text,
        reply_markup=admin_confirm_kb(user_id)
    )

    await message.answer(
        "Заявка отправлена!",
        reply_markup=main_kb
    )

    await state.clear()


# ---------- Подтверждение ----------

@dp.callback_query(F.data.startswith("confirm_"))
async def confirm(callback: CallbackQuery):
    user_id = int(callback.data.split("_")[1])

    update_status(user_id, "Подтвержден")

    await bot.send_message(
        user_id,
        "Запись подтверждена!"
    )

    await callback.message.edit_text("Подтверждено")


@dp.callback_query(F.data.startswith("reject_"))
async def reject(callback: CallbackQuery):
    user_id = int(callback.data.split("_")[1])

    update_status(user_id, "Отклонен")

    await bot.send_message(
        user_id,
        "Запись отклонена"
    )

    await callback.message.edit_text("Отклонено")


# ---------- Отмена ----------

@dp.message(F.text == "Отменить запись")
async def cancel(message: Message):
    update_status(message.from_user.id, "Отменил")
    await message.answer("Запись отменена")


# ---------- Активные ----------

@dp.message(F.text == "Активные записи")
async def active(message: Message):
    if message.from_user.id != ADMIN_ID:
        return

    users = get_active_users()

    if not users:
        await message.answer("Нет записей")
        return

    text = "Активные записи:\n\n"

    for row in users:
        text += (
            f"{row[3].value}\n"
            f"{row[5].value}\n\n"
        )

    await message.answer(text)


# ---------- Excel скачать ----------

@dp.message(F.text == "Скачать Excel")
async def send_excel(message: Message):
    if message.from_user.id != ADMIN_ID:
        return

    file = FSInputFile(EXCEL_FILE)

    await message.answer_document(file)


# ---------- Напоминания ----------

async def reminder_loop():
    while True:
        await asyncio.sleep(3600)

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        weekday = datetime.now().weekday()

        for row in ws.iter_rows(min_row=2):

            user_id = row[1].value
            status = row[6].value
            fri = row[7].value
            sat = row[8].value

            if status != "Подтвержден":
                continue

            try:

                if weekday == 4 and fri != "Да":

                    await bot.send_message(
                        user_id,
                        "Напоминание!\nЗанятие на выходных."
                    )

                    row[7].value = "Да"
                    wb.save(EXCEL_FILE)

                if weekday == 5 and sat != "Да":

                    await bot.send_message(
                        user_id,
                        "Напоминание!\nСегодня занятие."
                    )

                    row[8].value = "Да"
                    wb.save(EXCEL_FILE)

            except:
                pass


# ---------- Запуск ----------

async def main():
    init_excel()
    asyncio.create_task(reminder_loop())
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())